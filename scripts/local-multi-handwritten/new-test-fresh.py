import os
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font
from pdf2image import convert_from_path
from pytesseract import pytesseract
from textblob import TextBlob
from sklearn.feature_extraction.text import CountVectorizer

pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

def ocr_pdf(pdf_path):
    images = convert_from_path(pdf_path, dpi=300)
    text = ""
    for image in images:
        image = image.convert('L')  # Convert image to grayscale
        text += pytesseract.image_to_string(image, lang='eng', config='--psm 6')
    return text

pdf_folder = r'E:\handwritingtest-1\pdfs\Gun_Control_PDF\gunsedit'
pdf_texts = []
pdf_filenames = []

for pdf_file in os.listdir(pdf_folder):
    if pdf_file.endswith('.pdf'):
        full_path = os.path.join(pdf_folder, pdf_file)
        extracted_text = ocr_pdf(full_path)
        if extracted_text.strip():  # Only append if the text isn't empty
            pdf_texts.append(extracted_text)
            pdf_filenames.append(pdf_file)

csv_path = r'E:\handwritingtest-1\typewriting-test-1\predefined_data.csv'
df_csv = pd.read_csv(csv_path)
predefined_texts = df_csv['predefined text'].tolist()

def extract_keywords(text):
    try:
        vectorizer = CountVectorizer().fit([text])
        vectors = vectorizer.transform([text]).toarray()
        keywords = [word for word, count in zip(vectorizer.get_feature_names_out(), vectors[0]) if count > 2]
        return keywords
    except ValueError:
        return []

def match_text_with_predefined(text, predefined_texts):
    text_keywords = extract_keywords(text)
    matches = []
    for predefined_text in predefined_texts:
        words_in_predefined = set(predefined_text.split())
        common_words = set(text_keywords).intersection(words_in_predefined)

        if common_words:
            matches.append(predefined_text)
            if len(matches) == 3:
                break
    return matches

matched_texts = [match_text_with_predefined(text, predefined_texts) for text in pdf_texts]

def spellcheck(text):
    blob = TextBlob(text)
    corrected = blob.correct().string
    return corrected

corrected_texts = [spellcheck(text) for text in pdf_texts]

wb = Workbook()
ws = wb.active
ws.append(['Filename', 'Extracted Text', 'Match 1', 'Match 2', 'Match 3'])

for filename, corrected_text, matches in zip(pdf_filenames, corrected_texts, matched_texts):
    sanitized_text = corrected_text.replace('\f', '').strip()
    row = [filename, sanitized_text] + matches + [''] * (3 - len(matches))
    ws.append(row)

for row in ws.iter_rows(min_row=2, max_col=2, max_row=ws.max_row):
    for cell in row:
        blob = TextBlob(cell.value)
        if blob.words != blob.correct().words:
            cell.font = Font(color="FF0000")

wb.save(r'E:\handwritingtest-1\typewriting-test-1\results.xlsx')
